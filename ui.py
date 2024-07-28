import time
from copy import copy

from PyQt6.QtCore import Qt
from PyQt6.QtGui import QAction, QIcon, QCursor
from PyQt6.QtWidgets import QTableWidget, QTableWidgetItem, QMenu, QInputDialog, QComboBox, QFormLayout
from openpyxl import load_workbook, workbook
from unit_price import price_dict
import threading


class xlsx_ui(QTableWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.cellChanged["int", "int"].connect(self.update_cell)
        # 声明在groupBox创建右键菜单
        self.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.customContextMenuRequested.connect(self.rightEvent)  # 连接到菜单显示函数

        self.wb = workbook.Workbook()
        self.ws = self.wb.active

        self.ws.cell(1, 1, "商品")
        self.ws.cell(1, 2, "标价")
        self.ws.cell(1, 3, "售价")
        self.ws.cell(1, 4, "数量")
        self.ws.cell(1, 5, "标准总价")
        self.ws.cell(1, 6, "销售总价")

        self.running = True
        self.init()
        self.update()
        # self.thread1 = threading.Thread(target=self.run)
        # self.thread1.start()

    def run(self):
        while self.running:
            self.update()
            time.sleep(1)

    def setItemValue(self, row, col, sheet):
        value = sheet.cell(row, col).value
        if value is None:
            if col == 1:
                value = "未命名"
            else:
                value = 0
        self.setItem(row - 1, col - 1, QTableWidgetItem(str(value)))

    def init(self):
        self.blockSignals(True)
        # 读取Excel文件
        t_wb = load_workbook('default.xlsx')
        t_ws = t_wb.active

        for row in range(2, t_ws.max_row + 1):
            self.ws.cell(row, 1).value = t_ws.cell(row, 1).value
            self.ws.cell(row, 3).value = 0 if t_ws.cell(row, 2).value is None else float(t_ws.cell(row, 2).value)
            self.ws.cell(row, 4).value = 0 if t_ws.cell(row, 3).value is None else float(t_ws.cell(row, 3).value)
            # self.ws.cell(row, 2).value = price_dict[self.ws.cell(row, 1).value]
            # self.ws.cell(row, 5).value = int(self.ws.cell(row, 4).value) * float(self.ws.cell(row, 2).value)
            # self.ws.cell(row, 6).value = int(self.ws.cell(row, 4).value) * float(self.ws.cell(row, 3).value)
        self.blockSignals(False)

    def getIntValue(self, num):
        value = 0
        try:
            value = int(num)
        except Exception as e:
            value = 0
        return value

    def getFloatValue(self, num):
        value = 0.0
        try:
            value = float(num)
        except Exception as e:
            value = 0.0
        return value

    def update(self):
        self.blockSignals(True)
        self.clear()
        all_value, all_fix_value = 0.0, 0.0

        for row in range(2, self.ws.max_row + 1):
            self.ws.cell(row, 2).value = price_dict[self.ws.cell(row, 1).value] if self.ws.cell(row, 1).value in price_dict else 0.0
            self.ws.cell(row, 5).value = self.getIntValue(self.ws.cell(row, 4).value) * self.getFloatValue(self.ws.cell(row, 2).value)
            self.ws.cell(row, 6).value = self.getIntValue(self.ws.cell(row, 4).value) * self.getFloatValue(self.ws.cell(row, 3).value)
            all_value += self.ws.cell(row, 5).value
            all_fix_value += self.ws.cell(row, 6).value

        # 设置表格的行数和列数
        self.setRowCount(self.ws.max_row + 1)
        self.setColumnCount(self.ws.max_column)

        # 读取数据并显示在表格中
        for row in range(1, self.ws.max_row + 1):
            for col in range(1, self.ws.max_column + 1):
                self.setItemValue(row, col, self.ws)

        self.setItem(self.rowCount() - 1, self.columnCount() - 2, QTableWidgetItem(str(all_value)))
        self.setItem(self.rowCount() - 1, self.columnCount() - 1, QTableWidgetItem(str(all_fix_value)))
        self.blockSignals(False)

    def resizeEvent(self, e):
        if self.columnCount() < 3:
            return
        col_with = int(self.width() / self.columnCount() - 3)
        for index in range(self.columnCount()):
            self.setColumnWidth(index, col_with)

    def update_cell(self, row, column):
        if row == 0:
            return
        if column != 0:
            self.ws.cell(row + 1, column + 1, self.getFloatValue(self.item(row, column).text()))
        else:
            self.ws.cell(row + 1, column + 1, self.item(row, column).text())

        if (column == 1 and self.ws.cell(row + 1, column + 1).value > 1e-8 and self.item(row, 0) is not None
                and self.item(row, 0).text() not in price_dict and self.item(row, 0).text() != "" ):
            print("保存价格")
            price_dict[self.item(row, 0).text()] = self.ws.cell(row + 1, column + 1).value
        print(row, column, self.item(row, column).text())
        self.update()

    # 创建右键菜单函数
    def rightEvent(self, pos):
        row, col = 0, 0
        # 获取点击行号
        for i in self.selectionModel().selection().indexes():
            row = i.row()
            col = i.column()

        # 菜单对象
        menu = QMenu(self)

        actionSaveFile = QAction("保存数据", self)
        actionSaveFile.triggered.connect(self.SaveFile)  # 将动作A触发时连接到槽函数 button
        menu.addAction(actionSaveFile)

        actionAddRow = QAction("添加商品", self)
        actionAddRow.triggered.connect(self.AddRow)  # 将动作A触发时连接到槽函数 button
        menu.addAction(actionAddRow)

        actionAddRow = QAction("删除商品", self)
        actionAddRow.triggered.connect(lambda: self.DelRow(row))  # 将动作A触发时连接到槽函数 button
        menu.addAction(actionAddRow)

        if col == 1:
            actionBaseValue = QAction("修改标价", self)
            actionBaseValue.triggered.connect(lambda: self.changeBaseValue(row, col))  # 将动作A触发时连接到槽函数 button
            menu.addAction(actionBaseValue)

        if col == 2:
            actionMulBaseValue = QAction("标价折扣", self)
            actionMulBaseValue.triggered.connect(lambda: self.BasevalueStar(row, col))  # 将动作A触发时连接到槽函数 button
            menu.addAction(actionMulBaseValue)
            actionMulValue = QAction("售价折扣", self)
            actionMulValue.triggered.connect(lambda: self.valueStar(row, col))  # 将动作A触发时连接到槽函数 button
            menu.addAction(actionMulValue)

        screenPos = self.mapToGlobal(pos)
        action = menu.exec(screenPos)

    def SaveFile(self):
        save_wb = copy(self.wb)
        save_ws = save_wb.active

        all_value, all_fix_value = 0.0, 0.0
        for row in range(2, save_ws.max_row + 1):
            all_value += save_ws.cell(row=row, column=5).value
            all_fix_value += save_ws.cell(row=row, column=6).value
        save_ws.cell(save_ws.max_row + 1, 5, all_value)
        save_ws.cell(save_ws.max_row, 6, all_fix_value)
        save_wb.save('default_t.xlsx')

        price_dict_wb = workbook.Workbook()
        price_dict_ws = price_dict_wb.active
        for item in price_dict:
            price_dict_ws.cell(price_dict_ws.max_row + 1, 1).value = item
            price_dict_ws.cell(price_dict_ws.max_row, 2).value = price_dict[item]
        price_dict_wb.save("price.xlsx")


    def AddRow(self):
        text, okPressed = QInputDialog.getText(self, "输入商品名", "商品名")
        self.ws.cell(self.ws.max_row + 1, 1, text)
        self.ws.cell(self.ws.max_row, 2, 0.0)
        self.ws.cell(self.ws.max_row, 3, 0.0)
        self.ws.cell(self.ws.max_row, 4, 0.0)
        self.ws.cell(self.ws.max_row, 5, 0.0)
        self.ws.cell(self.ws.max_row, 6, 0.0)
        self.update()

    def DelRow(self, row):
        self.ws.delete_rows(row + 1)
        self.update()

    def changeBaseValue(self, row, col):
        value, okPressed = QInputDialog.getDouble(self, "输入价格", "标价")
        if self.ws.cell(row + 1, 1).value in price_dict:
            price_dict[self.ws.cell(row + 1, 1).value] = value
        self.update()

    def BasevalueStar(self, row, col):
        mul, okPressed = QInputDialog.getDouble(self, "输入折扣", "标价折扣")
        if okPressed:
            self.ws.cell(row + 1, col + 1).value = mul * self.ws.cell(row + 1, col).value
        self.update()

    def valueStar(self, row, col):
        mul, okPressed = QInputDialog.getDouble(self, "输入折扣", "售价折扣")
        if okPressed:
            self.ws.cell(row + 1, col + 1).value *= mul
        self.update()