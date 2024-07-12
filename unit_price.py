import openpyxl

price_dict = {}


def update_price():
    price_file = openpyxl.load_workbook("price.xlsx")
    sheet = price_file.active
    # 获取行数和列数
    rows = sheet.max_row
    # 读取数据并显示在表格中
    for row in range(2, rows):
        item = sheet.cell(row=row, column=1).value
        value = sheet.cell(row=row, column=2).value
        if item is not None and value is not None:
            price_dict[item] = value


update_price()
