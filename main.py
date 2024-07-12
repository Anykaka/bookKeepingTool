import sys
from PyQt6.QtWidgets import QApplication, QMainWindow
from openpyxl import load_workbook
import unit_price, xlsx_ui, operate_ui


class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.initUI()

    def initUI(self):
        self.tableWidget = QTableWidget(self)
        self.setCentralWidget(self.tableWidget)

        # 读取Excel文件
        wb = load_workbook('default.xlsx')
        sheet = wb.active

        # 获取行数和列数
        rows = sheet.max_row
        cols = sheet.max_column

        # 设置表格的行数和列数
        self.tableWidget.setRowCount(rows)
        self.tableWidget.setColumnCount(cols)

        # 读取数据并显示在表格中
        for row in range(1, rows + 1):
            for col in range(1, cols + 1):
                cell_value = sheet.cell(row=row, column=col).value
                item = QTableWidgetItem(str(cell_value))
                self.tableWidget.setItem(row - 1, col - 1, item)

        self.show()


if __name__ == '__main__':
    print(unit_price.price_dict)
    app = QApplication(sys.argv)
    mainWindow = MainWindow()
    sys.exit(app.exec())
